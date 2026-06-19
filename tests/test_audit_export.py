"""PRD 0001 — audit-log export round-trip.

Record known audit rows → export → parse BOTH the CSV (stdlib csv) and the XLSX
(openpyxl) back → assert the parsed rows match the recorded rows (count, fields,
order). Also pins the export tool as a plain READ (not write / not gated).
"""

import base64
import csv
import io

import pytest
from openpyxl import load_workbook

from app import audit_export
from agent import audit, config
from agent import tools as T


@pytest.fixture(autouse=True)
def _tmp_audit(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "AUDIT_DB", tmp_path / "audit.sqlite")
    yield


# Three known rows, recorded oldest → newest; covers a None field (old_verdict).
_RECORDS = [
    ("agent-user", "override", "r1", "FLAG", "PASS", "manual review confirms compliant"),
    ("agent-user", "manual_entry", "r2", None, "abv=5.0", "OCR could not read ABV"),
    ("agent-user", "override", "r3", "PASS", "FLAG", "warning header is title-case"),
]


def _seed():
    for rec in _RECORDS:
        audit.record(*rec)
    return audit.all_rows()


def _assert_matches(parsed, rows):
    """Parsed (str) cells equal the recorded rows in count, fields, and order."""
    assert len(parsed) == len(rows)
    for got, row in zip(parsed, rows):
        assert list(got.keys()) == audit_export.COLUMNS          # exact column order
        for col in audit_export.COLUMNS:
            expected = "" if row.get(col) is None else str(row[col])
            assert got[col] == expected


def test_csv_round_trip():
    rows = _seed()
    parsed = list(csv.DictReader(io.StringIO(audit_export.audit_to_csv(rows))))
    _assert_matches(parsed, rows)


def test_xlsx_round_trip():
    rows = _seed()
    wb = load_workbook(io.BytesIO(audit_export.audit_to_xlsx(rows)))
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    header = list(values[0])
    assert header == audit_export.COLUMNS
    parsed = [dict(zip(header, ["" if v is None else str(v) for v in r]))
              for r in values[1:]]
    _assert_matches(parsed, rows)


def test_export_tool_returns_both_formats_matching_helpers():
    rows = _seed()
    out = T.export_audit_log.invoke({})
    assert out["count"] == len(_RECORDS)
    by_name = {d["filename"]: d for d in out["downloads"]}
    assert set(by_name) == {"audit_log.csv", "audit_log.xlsx"}
    # The tool's payloads are exactly the export helpers' output over all_rows().
    assert base64.b64decode(by_name["audit_log.csv"]["b64"]).decode() == \
        audit_export.audit_to_csv(rows)
    csv_parsed = list(csv.DictReader(io.StringIO(
        base64.b64decode(by_name["audit_log.csv"]["b64"]).decode())))
    _assert_matches(csv_parsed, rows)


def test_export_empty_log_has_header_only():
    out = T.export_audit_log.invoke({})
    assert out["count"] == 0
    by_name = {d["filename"]: d for d in out["downloads"]}
    csv_text = base64.b64decode(by_name["audit_log.csv"]["b64"]).decode()
    assert list(csv.reader(io.StringIO(csv_text))) == [audit_export.COLUMNS]


def test_formula_injection_in_reason_is_neutralized():
    # A reason beginning with a formula trigger must be defused with a leading
    # apostrophe in BOTH formats so it can't execute when opened in a spreadsheet.
    payload = "=cmd|'/c calc'!A1"
    audit.record("agent-user", "override", "r9", "FLAG", "PASS", payload)
    rows = audit.all_rows()

    csv_rows = list(csv.DictReader(io.StringIO(audit_export.audit_to_csv(rows))))
    assert csv_rows[-1]["reason"] == "'" + payload          # CSV defused
    assert csv_rows[-1]["actor"] == "agent-user"            # benign cell untouched

    wb = load_workbook(io.BytesIO(audit_export.audit_to_xlsx(rows)))
    last = list(wb.active.iter_rows(values_only=True))[-1]
    assert last[audit_export.COLUMNS.index("reason")] == "'" + payload   # XLSX identical


@pytest.mark.parametrize("trigger", ["=", "+", "-", "@"])
def test_all_formula_triggers_defused(trigger):
    audit.record("agent-user", "override", "rX", "FLAG", "PASS", trigger + "danger")
    csv_rows = list(csv.DictReader(io.StringIO(
        audit_export.audit_to_csv(audit.all_rows()))))
    assert csv_rows[-1]["reason"] == "'" + trigger + "danger"


def test_export_tool_is_plain_read_not_gated():
    names = {t.name for t in T.READ_TOOLS}
    assert "export_audit_log" in names
    assert "export_audit_log" not in T.WRITE_TOOL_NAMES
    assert "export_audit_log" not in {t.name for t in T.WRITE_TOOLS}
